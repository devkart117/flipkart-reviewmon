import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


def get_google_sheet_as_dataframe(url):
    df = pd.read_csv(url)
    return df


def get_flipkart_data():
    return get_google_sheet_as_dataframe('https://docs.google.com/spreadsheets/d/1kgqFc2pYijdyYhcDB0JJ3dwnY5ONj5zMlb576hdaZj0/export?gid=31823053&format=csv').to_dict(orient='records')


def save_data(data_list):
    df = pd.DataFrame(data_list)

    # Create a Pandas Excel writer using openpyxl as the engine
    with pd.ExcelWriter('data/output.xlsx', engine='openpyxl') as writer:
        # Convert the dataframe to an Excel object
        df.to_excel(writer, index=False, sheet_name='Sheet1')

        # Access the openpyxl workbook and sheet objects
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # Define a light gray fill
        fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

        # Apply the fill to the specific columns (3rd, 5th, 7th, 9th, 11th)
        for col in [4, 6, 8, 10, 12, 14]:
            for row in range(1, len(df) + 2):  # +2 to account for 1-based index in openpyxl and header
                cell = worksheet.cell(row=row, column=col)
                cell.fill = fill

        # Adjust column widths based on the maximum length of the values in each column
        for col_num in range(1, len(df.columns) + 1):  # Loop through all columns
            max_length = 0
            column = df.iloc[:, col_num - 1]  # Get the data of the column
            
            # Check the maximum length of the values in the column
            for value in column:
                max_length = max(max_length, len(str(value)))
            
            # Also consider the length of the header (the column name)
            header_length = len(df.columns[col_num - 1])
            max_length = max(max_length, header_length)
            
            # Set the width to max_length + 2
            worksheet.column_dimensions[chr(64 + col_num)].width = max_length + 2

        # Save the workbook
        workbook.save('data/output.xlsx')